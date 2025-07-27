from datetime import datetime, timedelta
import os
from flask import Blueprint, render_template, redirect, url_for, request, flash, jsonify, current_app
from flask_login import login_required, current_user
from functools import wraps
from models import User, Role, SubscriptionPlan, Subscription, Classroom, Notification, Payment, SystemSettings, SubscriptionPayment, ClassroomEnrollment, Banner
from flask_wtf.csrf import CSRFProtect, CSRFError
from models import db


# استيراد db من داخل الدالة لتجنب الاستيراد الدائري
def get_db():
    global db
    if not hasattr(db, 'session'):
        from app import db as app_db
        return app_db
    return db







admin_bp = Blueprint('admin', __name__)

# CSRF error handler
@admin_bp.errorhandler(CSRFError)
def handle_csrf_error(e):
    flash('انتهت صلاحية النموذج. يرجى المحاولة مرة أخرى.', 'danger')
    return redirect(url_for('admin.subscriptions'))


# Admin required decorator
def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated or current_user.role != Role.ADMIN:
            flash('غير مصرح لك بالوصول إلى هذه الصفحة', 'danger')
            return redirect(url_for('main.index'))
        return f(*args, **kwargs)
    return decorated_function

# دالة للتحقق من نوع الجهاز (موبايل أو جهاز مكتبي)
def is_mobile():
    user_agent = request.headers.get('User-Agent', '').lower()
    mobile_patterns = [
        'android', 'iphone', 'ipod', 'windows phone', 'mobile', 'tablet',
        'blackberry', 'opera mini', 'opera mobi', 'webos', 'fennec'
    ]
    return any(pattern in user_agent for pattern in mobile_patterns)

@admin_bp.route('/dashboard')
@login_required
@admin_required
def dashboard():
    # Get counts for dashboard statistics
    users_count = User.query.count()
    teachers_count = User.query.filter_by(role=Role.TEACHER).count()
    students_count = User.query.filter_by(role=Role.STUDENT).count()
    classrooms_count = Classroom.query.count()
    active_subscriptions = Subscription.query.filter(Subscription.end_date > datetime.utcnow()).count()

    # Calculate monthly revenue from subscription payments only
    start_of_month = datetime.utcnow().replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    end_of_month = (start_of_month + timedelta(days=32)).replace(day=1) - timedelta(seconds=1)

    # Get all approved subscription payments for current month
    monthly_revenue = db.session.query(db.func.sum(SubscriptionPayment.amount))\
        .filter(SubscriptionPayment.created_at.between(start_of_month, end_of_month))\
        .filter(SubscriptionPayment.status == 'approved')\
        .scalar() or 0

    # Get recent users
    recent_users = User.query.order_by(User.created_at.desc()).limit(10).all()

    # Sample data for enrollment chart
    now = datetime.utcnow()
    enrollment_dates = [(now - timedelta(days=i)).strftime('%Y-%m-%d') for i in range(30, 0, -1)]
    enrollment_counts = [0] * 30  # Will be populated with actual data in a real implementation

    template = 'admin/admin-mobile/dashboard.html' if is_mobile() else 'admin/dashboard.html'

    # الحصول على قيم الألوان من إعدادات النظام
    primary_color = SystemSettings.get_setting('primary_color', '#3498db')  # اللون الافتراضي
    secondary_color = SystemSettings.get_setting('secondary_color', '#2ecc71')  # اللون الافتراضي


    return render_template(template,
                        user_count=users_count,
                        primary_color=primary_color,
                        secondary_color=secondary_color,
                        teacher_count=teachers_count,
                        student_count=students_count,
                        classroom_count=classrooms_count,
                        subscription_count=active_subscriptions,
                        recent_users=recent_users,
                        enrollment_dates=enrollment_dates,
                        enrollment_counts=enrollment_counts,
                        revenue=monthly_revenue)

@admin_bp.route('/users', methods=['GET', 'POST'])
@login_required
@admin_required
def users():

    
    # إذا كان طلب POST، فقم بمعالجة تحديث أو إضافة المستخدم
    if request.method == 'POST':
        action = request.form.get('action')

        if action == 'update':
            user_id = request.form.get('user_id', type=int)
            if user_id:
                user = User.query.get_or_404(user_id)
                user.name = request.form.get('name')
                user.email = request.form.get('email')
                user.phone = request.form.get('phone')
                user.role = request.form.get('role')

                db.session.commit()
                flash('تم تحديث بيانات المستخدم بنجاح', 'success')
                return redirect(url_for('admin.users'))

        elif action == 'add':
            # إضافة مستخدم جديد
            from werkzeug.security import generate_password_hash

            name = request.form.get('name')
            email = request.form.get('email')
            phone = request.form.get('phone')
            role = request.form.get('role')
            password = request.form.get('password')

            # التحقق من البريد الإلكتروني
            existing_email = User.query.filter_by(email=email).first()
            if existing_email:
                flash('البريد الإلكتروني مسجل بالفعل', 'danger')
                return redirect(url_for('admin.users'))

            # إنشاء مستخدم جديد
            new_user = User(
                name=name,
                email=email,
                phone=phone,
                role=role,
                is_active=True
            )
            new_user.set_password(password)

            db.session.add(new_user)
            db.session.commit()

            flash('تمت إضافة المستخدم بنجاح', 'success')
            return redirect(url_for('admin.users'))

    # Get filter parameters
    role_filter = request.args.get('role', '')
    status_filter = request.args.get('status', '')
    search = request.args.get('search', '')

    # Base query
    query = User.query

    # Apply filters
    if role_filter and role_filter in [Role.STUDENT, Role.TEACHER, Role.ASSISTANT, Role.ADMIN]:
        query = query.filter_by(role=role_filter)

    if status_filter:
        if status_filter == 'active':
            query = query.filter_by(is_active=True)
        elif status_filter == 'inactive':
            query = query.filter_by(is_active=False)

    if search:
        query = query.filter(
            (User.name.ilike(f'%{search}%')) | 
            (User.phone.ilike(f'%{search}%')) |
            (User.email.ilike(f'%{search}%'))
        )

    # Get paginated results
    page = request.args.get('page', 1, type=int)
    users = query.order_by(User.created_at.desc()).paginate(page=page, per_page=20)

    template = 'admin/admin-mobile/users.html' if is_mobile() else 'admin/users.html'

    # الحصول على قيم الألوان من إعدادات النظام
    primary_color = SystemSettings.get_setting('primary_color', '#3498db')  # اللون الافتراضي
    secondary_color = SystemSettings.get_setting('secondary_color', '#2ecc71')  # اللون الافتراضي

    


    return render_template(template, 
                           users=users, 
                           role=role_filter, 
                           status=status_filter, 
                           search=search, 
                           current_time=datetime.now(),
                           primary_color=primary_color,
                           secondary_color=secondary_color,)


@admin_bp.route('/user/<int:user_id>/toggle_status', methods=['POST'])
@login_required
@admin_required
def toggle_user_status(user_id):
    user = User.query.get_or_404(user_id)
    response = {'success': False}

    # التحقق مما إذا كان الطلب AJAX
    is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'

    # Don't allow deactivating the current admin
    if user.id == current_user.id:
        message = 'لا يمكنك تعطيل حسابك الخاص'
        flash(message, 'danger')
        response['message'] = message
        return jsonify(response) if is_ajax else redirect(url_for('admin.users'))

    # Don't allow non-admin users to modify admin accounts
    if user.role == Role.ADMIN and current_user.role != Role.ADMIN:
        message = 'لا يمكنك تعديل حسابات المشرفين'
        flash(message, 'danger')
        response['message'] = message
        return jsonify(response) if is_ajax else redirect(url_for('admin.users'))

    try:
        # تبديل حالة المستخدم
        user.is_active = not user.is_active
        db.session.commit()

        # إرسال رسالة نجاح مناسبة
        status = "تفعيل" if user.is_active else "تعطيل"
        message = f'تم {status} حساب {user.name} بنجاح'
        flash(message, 'success')
        
        response['success'] = True
        response['message'] = message
        response['is_active'] = user.is_active

    except Exception as e:
        db.session.rollback()
        message = 'حدث خطأ أثناء تحديث حالة المستخدم'
        flash(message, 'danger')
        response['message'] = message
        
    return jsonify(response) if is_ajax else redirect(url_for('admin.users'))


@admin_bp.route('/user/<int:user_id>/reset_password', methods=['POST'])
@login_required
@admin_required
def reset_user_password(user_id):
    user = User.query.get_or_404(user_id)

    # إعادة تعيين كلمة المرور لكلمة المرور الافتراضية
    default_password = "12345678"
    user.set_password(default_password)
    
    db.session.commit()

    flash(f'تم إعادة تعيين كلمة المرور للمستخدم {user.name} إلى: {default_password}', 'success')
    return redirect(url_for('admin.users'))


@admin_bp.route('/subscriptions')
@login_required
@admin_required
def subscriptions():
    # Get filter parameters
    status = request.args.get('status', '')
    plan_id = request.args.get('plan_id', type=int)
    search = request.args.get('search', '')
    
    # Get subscription plans
    plans = SubscriptionPlan.query.all()

    # Get active subscriptions
    active_subs = Subscription.query.filter(Subscription.end_date > datetime.utcnow()).order_by(Subscription.start_date.desc()).all()

    # Get all subscriptions for the comprehensive table with pagination
    page = request.args.get('page', 1, type=int)
    subscriptions = Subscription.query.order_by(Subscription.start_date.desc()).paginate(page=page, per_page=10)

    # Get all subscription payments
    payments = SubscriptionPayment.query.order_by(SubscriptionPayment.created_at.desc()).all()

    # حساب إجمالي الإيرادات
    total_revenue = 0
    if payments:
        for payment in payments:
            if payment.status == 'approved' and payment.amount:
                total_revenue += payment.amount

    # Calculate monthly revenue
    start_of_month = datetime.utcnow().replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    end_of_month = (start_of_month + timedelta(days=32)).replace(day=1) - timedelta(seconds=1)

    # Get all approved subscription payments for current month
    month_revenue = db.session.query(db.func.sum(SubscriptionPayment.amount))\
        .filter(SubscriptionPayment.created_at.between(start_of_month, end_of_month))\
        .filter(SubscriptionPayment.status == 'approved')\
        .scalar() or 0

    # Get current time for template
    now = datetime.utcnow()
    current_time = now  # نسخة من الوقت الحالي لاستخدامها في القالب

    # قم بحساب عدد المدفوعات قيد الانتظار
    pending_payments = SubscriptionPayment.query.filter_by(status='pending').count()

    # جلب طلبات الاشتراك المعلقة
    pending_subscription_payments = SubscriptionPayment.query.filter_by(status='pending')\
        .order_by(SubscriptionPayment.created_at.desc())\
        .all()

    template = 'admin/admin-mobile/subscriptions.html' if is_mobile() else 'admin/subscriptions.html'

    # الحصول على قيم الألوان من إعدادات النظام
    primary_color = SystemSettings.get_setting('primary_color', '#3498db')  # اللون الافتراضي
    secondary_color = SystemSettings.get_setting('secondary_color', '#2ecc71')  # اللون الافتراضي


    return render_template(template, 
                         plans=plans,
                         primary_color=primary_color,
                         secondary_color=secondary_color, 
                         active_subs=active_subs,
                         pending_subscription_payments=pending_subscription_payments, 
                         subscriptions=subscriptions, 
                         all_subscriptions=Subscription.query.order_by(Subscription.start_date.desc()).all(),
                         payments=payments,
                         now=now,
                         current_time=now,  # إضافة متغير current_time ليكون متاحًا في القالب
                         total_revenue=total_revenue,
                         month_revenue=month_revenue,
                         pending_payments=pending_payments,
                         subscription_plans=plans,
                         status=status,
                         plan_id=plan_id,
                         search=search)


@admin_bp.route('/subscription_payment/<int:payment_id>/approve', methods=['POST'])
@login_required
@admin_required
def approve_subscription_payment(payment_id):
    payment = SubscriptionPayment.query.get_or_404(payment_id)
    
    if payment.status != 'pending':
        flash('لا يمكن قبول هذا الطلب لأنه تم معالجته مسبقاً', 'warning')
        return redirect(url_for('admin.subscriptions'))

    # إنشاء اشتراك جديد
    subscription = Subscription(
        user_id=payment.user_id,
        plan_id=payment.plan_id,
        start_date=datetime.utcnow(),
        end_date=datetime.utcnow() + timedelta(days=payment.plan.duration_days),
        is_active=True
    )

    # تحديث حالة الدفع
    payment.status = 'approved'
    payment.processed_at = datetime.utcnow()

    db.session.add(subscription)
    db.session.commit()

    flash('تم قبول طلب الاشتراك بنجاح', 'success')
    return redirect(url_for('admin.subscriptions'))



@admin_bp.route('/subscription_payment/<int:payment_id>/reject', methods=['POST'])
@login_required
@admin_required
def reject_subscription_payment(payment_id):
    payment = SubscriptionPayment.query.get_or_404(payment_id)
    
    if payment.status != 'pending':
        flash('لا يمكن رفض هذا الطلب لأنه تم معالجته مسبقاً', 'warning')
        return redirect(url_for('admin.subscriptions'))

    # تحديث حالة الدفع
    payment.status = 'rejected'
    payment.processed_at = datetime.utcnow()
    db.session.commit()

    flash('تم رفض طلب الاشتراك', 'success')
    return redirect(url_for('admin.subscriptions'))

@admin_bp.route('/subscription_plan/new', methods=['GET', 'POST'])
@login_required
@admin_required
def new_subscription_plan():
    if request.method == 'POST':
        # Validate CSRF token
        if not request.form.get('csrf_token'):
            flash('خطأ: رمز الحماية CSRF غير موجود', 'danger')
            return redirect(url_for('admin.subscriptions'))

        name = request.form.get('name')
        description = request.form.get('description')
        price = float(request.form.get('price', 0))
        duration_days = int(request.form.get('duration_days', 30))
        max_classrooms = int(request.form.get('max_classrooms', 1))
        has_chat = 'has_chat' in request.form
        allow_assistant = 'allow_assistant' in request.form
        advanced_analytics = 'advanced_analytics' in request.form

        new_plan = SubscriptionPlan(
            name=name,
            description=description,
            price=price,
            duration_days=duration_days,
            max_classrooms=max_classrooms,
            has_chat=has_chat,
            allow_assistant=allow_assistant,
            advanced_analytics=advanced_analytics
        )

        db.session.add(new_plan)
        db.session.commit()

        flash('تم إنشاء باقة الاشتراك بنجاح', 'success')
        return redirect(url_for('admin.subscriptions'))
    
    # الحصول على قيم الألوان من إعدادات النظام
    primary_color = SystemSettings.get_setting('primary_color', '#3498db')  # اللون الافتراضي
    secondary_color = SystemSettings.get_setting('secondary_color', '#2ecc71')  # اللون الافتراضي


    return render_template('admin/edit_subscription_plan.html',
                           primary_color=primary_color,
                           secondary_color=secondary_color,)

@admin_bp.route('/subscription_plan/<int:plan_id>/edit', methods=['GET', 'POST'])
@login_required
@admin_required
def edit_subscription_plan(plan_id):
    plan = SubscriptionPlan.query.get_or_404(plan_id)

    if request.method == 'POST':
        # Validate CSRF token
        if not request.form.get('csrf_token'):
            flash('خطأ: رمز الحماية CSRF غير موجود', 'danger')
            return redirect(url_for('admin.subscriptions'))

        plan.name = request.form.get('name')
        plan.description = request.form.get('description')
        plan.price = float(request.form.get('price', 0))
        plan.duration_days = int(request.form.get('duration_days', 30))
        plan.max_classrooms = int(request.form.get('max_classrooms', 1))
        plan.has_chat = 'has_chat' in request.form
        plan.allow_assistant = 'allow_assistant' in request.form
        plan.advanced_analytics = 'advanced_analytics' in request.form

        db.session.commit()

        flash('تم تحديث باقة الاشتراك بنجاح', 'success')
        return redirect(url_for('admin.subscriptions'))
    
    # الحصول على قيم الألوان من إعدادات النظام
    primary_color = SystemSettings.get_setting('primary_color', '#3498db')  # اللون الافتراضي
    secondary_color = SystemSettings.get_setting('secondary_color', '#2ecc71')  # اللون الافتراضي


    return render_template('admin/edit_subscription_plan.html', plan=plan,
                           primary_color=primary_color,
                           secondary_color=secondary_color,)

@admin_bp.route('/subscription_plan/<int:plan_id>/delete', methods=['POST'])
@login_required
@admin_required
def delete_subscription_plan(plan_id):
    plan = SubscriptionPlan.query.get_or_404(plan_id)

    # Check if plan is in use
    if plan.subscriptions:
        flash('لا يمكن حذف الباقة لأن هناك مستخدمين مشتركين بها', 'danger')
        return redirect(url_for('admin.subscriptions'))

    db.session.delete(plan)
    db.session.commit()

    flash('تم حذف باقة الاشتراك بنجاح', 'success')
    return redirect(url_for('admin.subscriptions'))

@admin_bp.route('/subscription/<int:subscription_id>/cancel', methods=['POST'])
@login_required
@admin_required
def cancel_subscription(subscription_id):
    subscription = Subscription.query.get_or_404(subscription_id)

    # Set end date to now to effectively cancel the subscription
    subscription.end_date = datetime.utcnow()
    subscription.is_active = False
    db.session.commit()

    flash(f'تم إلغاء اشتراك {subscription.user.name} في باقة {subscription.plan.name} بنجاح', 'success')
    return redirect(url_for('admin.subscriptions'))

@admin_bp.route('/subscription/cancel_form', methods=['POST'])
@login_required
@admin_required
def cancel_subscription_form():
    subscription_id = request.form.get('subscription_id')
    if not subscription_id:
        flash('معرف الاشتراك مطلوب', 'error')
        return redirect(url_for('admin.subscriptions'))
    
    subscription = Subscription.query.get_or_404(int(subscription_id))

    # Set end date to now to effectively cancel the subscription
    subscription.end_date = datetime.utcnow()
    subscription.is_active = False
    db.session.commit()

    flash(f'تم إلغاء اشتراك {subscription.user.name} في باقة {subscription.plan.name} بنجاح', 'success')
    return redirect(url_for('admin.subscriptions'))

@admin_bp.route('/subscription/delete', methods=['POST'])
@login_required
@admin_required
def delete_subscription():
    subscription_id = request.form.get('subscription_id', type=int)
    if not subscription_id:
        flash('معرف الاشتراك غير صالح', 'danger')
        return redirect(url_for('admin.subscriptions'))

    subscription = Subscription.query.get_or_404(subscription_id)

    # Save the subscription info for flash message
    user_name = subscription.user.name
    plan_name = subscription.plan.name

    # Delete the subscription
    db.session.delete(subscription)
    db.session.commit()

    flash(f'تم حذف اشتراك {user_name} في باقة {plan_name} بنجاح', 'success')
    return redirect(url_for('admin.subscriptions'))

@admin_bp.route('/subscription/<int:subscription_id>/extend', methods=['POST'])
@login_required
@admin_required
def extend_subscription(subscription_id):
    subscription = Subscription.query.get_or_404(subscription_id)
    
    # الحصول على عدد الأيام المراد إضافتها من النموذج
    days = request.form.get('days', type=int, default=30)
    note = request.form.get('note', '')
    
    if days <= 0:
        flash('يجب أن يكون عدد الأيام أكبر من صفر', 'danger')
        return redirect(url_for('admin.subscriptions'))
    
    # إذا كان الاشتراك منتهي، يتم تعيين تاريخ البدء الجديد من الآن
    if subscription.end_date < datetime.utcnow():
        subscription.start_date = datetime.utcnow()
        subscription.end_date = datetime.utcnow() + timedelta(days=days)
    else:
        # إذا كان الاشتراك نشطًا، يتم إضافة الأيام إلى تاريخ الانتهاء الحالي
        subscription.end_date = subscription.end_date + timedelta(days=days)
    
    # تنشيط الاشتراك
    subscription.is_active = True
    
    # إضافة ملاحظة للإدارة (يمكن إضافة حقل في النموذج)
    # يمكن إضافة حقل notes في جدول الاشتراكات في المستقبل
    
    db.session.commit()
    
    flash(f'تم تمديد اشتراك {subscription.user.name} بنجاح لمدة {days} يوم', 'success')
    return redirect(url_for('admin.subscriptions'))

@admin_bp.route('/assign_trial/<int:user_id>', methods=['POST'])
@login_required
@admin_required
def assign_trial(user_id):
    user = User.query.get_or_404(user_id)

    # Only assign trial to teachers
    if user.role != Role.TEACHER:
        flash('يمكن تعيين فترة تجريبية للمعلمين فقط', 'danger')
        return redirect(url_for('admin.users'))

    # Get the premium plan
    premium_plan = SubscriptionPlan.query.order_by(SubscriptionPlan.price.desc()).first()

    if not premium_plan:
        flash('لا توجد باقات اشتراك بعد', 'danger')
        return redirect(url_for('admin.subscriptions'))

    # Check if user already has an active subscription
    active_sub = Subscription.query.filter(
        Subscription.user_id == user.id,
        Subscription.end_date > datetime.utcnow()
    ).first()

    if active_sub:
        flash('المستخدم لديه اشتراك فعال بالفعل', 'warning')
        return redirect(url_for('admin.users'))

    # Create trial subscription
    trial_days = int(request.form.get('trial_days', 14))
    trial_subscription = Subscription(
        user_id=user.id,
        plan_id=premium_plan.id,
        start_date=datetime.utcnow(),
        end_date=datetime.utcnow() + timedelta(days=trial_days),
        is_active=True,
        is_trial=True
    )

    db.session.add(trial_subscription)
    db.session.commit()

    flash(f'تم تعيين فترة تجريبية لمدة {trial_days} يوم للمستخدم بنجاح', 'success')
    return redirect(url_for('admin.users'))

@admin_bp.route('/notifications')
@login_required
@admin_required
def notifications():
    # Get recent notifications
    recent_notifications = Notification.query.order_by(Notification.created_at.desc()).limit(50).all()

    # Get all users for the notification form
    users = User.query.order_by(User.name).all()

    # Get current time for template
    now = datetime.utcnow()

    template = 'admin/admin-mobile/notifications.html' if is_mobile() else 'admin/notifications.html'

    # الحصول على قيم الألوان من إعدادات النظام
    primary_color = SystemSettings.get_setting('primary_color', '#3498db')  # اللون الافتراضي
    secondary_color = SystemSettings.get_setting('secondary_color', '#2ecc71')  # اللون الافتراضي


    return render_template(template, 
                           notifications=recent_notifications, 
                           users=users, now=now,
                           primary_color=primary_color,
                           secondary_color=secondary_color,)

@admin_bp.route('/send_notification', methods=['POST'])
@login_required
@admin_required
def send_notification():
    title = request.form.get('title')
    message = request.form.get('message')
    recipient_type = request.form.get('recipient_type')
    user_id = request.form.get('user_id')

    if not title or not message:
        flash('الرجاء إدخال العنوان والرسالة', 'danger')
        return redirect(url_for('admin.notifications'))

    notifications = []

    if recipient_type == 'all':
        # Send to all users
        users = User.query.all()
        for user in users:
            notification = Notification(
                user_id=user.id,
                title=title,
                message=message
            )
            notifications.append(notification)

    elif recipient_type == 'role':
        role = request.form.get('role')
        if role in [Role.STUDENT, Role.TEACHER, Role.ASSISTANT, Role.ADMIN]:
            # Send to users with specific role
            users = User.query.filter_by(role=role).all()
            for user in users:
                notification = Notification(
                    user_id=user.id,
                    title=title,
                    message=message
                )
                notifications.append(notification)

    elif recipient_type == 'user':
        # Send to specific user
        if user_id:
            user = User.query.get(user_id)
            if user:
                notification = Notification(
                    user_id=user.id,
                    title=title,
                    message=message
                )
                notifications.append(notification)

    elif recipient_type == 'phone':
        phone = request.form.get('phone')
        if phone:
            user = User.query.filter_by(phone=phone).first()
            if user:
                notification = Notification(
                    user_id=user.id,
                    title=title,
                    message=message
                )
                notifications.append(notification)

    if notifications:
        db.session.add_all(notifications)
        db.session.commit()
        flash(f'تم إرسال {len(notifications)} إشعار بنجاح', 'success')
    else:
        flash('لم يتم العثور على مستخدمين للإرسال إليهم', 'warning')

    return redirect(url_for('admin.notifications'))

@admin_bp.route('/classrooms')
@login_required
@admin_required
def classrooms():
    # Get filter parameters
    teacher_id = request.args.get('teacher_id', type=int)
    is_free = request.args.get('is_free')
    search = request.args.get('search', '')

    # Get counts for stats
    student_count = User.query.filter_by(role=Role.STUDENT).count()
    teacher_count = User.query.filter_by(role=Role.TEACHER).count()

    # Base query
    query = Classroom.query

    # Apply filters
    if teacher_id:
        query = query.filter_by(teacher_id=teacher_id)

    if is_free is not None:
        is_free_bool = (is_free == 'true')
        query = query.filter_by(is_free=is_free_bool)

    if search:
        query = query.filter(
            (Classroom.name.ilike(f'%{search}%')) | 
            (Classroom.code.ilike(f'%{search}%')) |
            (Classroom.subject.ilike(f'%{search}%'))
        )

    # Get paginated results
    page = request.args.get('page', 1, type=int)
    pagination = query.order_by(Classroom.created_at.desc()).paginate(page=page, per_page=20)
    classrooms = pagination

    # Get all teachers for the filter dropdown
    teachers = User.query.filter_by(role=Role.TEACHER).all()

    template = 'admin/admin-mobile/classrooms.html' if is_mobile() else 'admin/classrooms.html'

    # الحصول على قيم الألوان من إعدادات النظام
    primary_color = SystemSettings.get_setting('primary_color', '#3498db')  # اللون الافتراضي
    secondary_color = SystemSettings.get_setting('secondary_color', '#2ecc71')  # اللون الافتراضي

    currency = SystemSettings.get_setting('default_currency' , 'جنية')


    return render_template(template, 
                         classrooms=classrooms,
                         primary_color=primary_color,
                         secondary_color=secondary_color,
                         teachers=teachers, 
                         currency=currency,
                         pagination=pagination,
                         teacher_id=teacher_id or 0,  # توفير قيمة افتراضية
                         is_free=is_free,
                         search=search,
                         student_count=student_count,
                         teacher_count=teacher_count)

@admin_bp.route('/classroom/<int:classroom_id>/confirm_delete')
@login_required
@admin_required
def confirm_delete_classroom(classroom_id):
    classroom = Classroom.query.get_or_404(classroom_id)
    return render_template('admin/delete_classroom_confirm.html', classroom=classroom)

@admin_bp.route('/classroom/delete', methods=['POST'])
@login_required
@admin_required
def delete_classroom():
    classroom_id = request.form.get('classroom_id', type=int)
    if not classroom_id:
        flash('معرف الفصل غير صالح', 'danger')
        return redirect(url_for('admin.classrooms'))

    classroom = Classroom.query.get_or_404(classroom_id)

    # Save the classroom name for the flash message
    classroom_name = classroom.name

    # Delete the classroom (will cascade delete all related records)
    db.session.delete(classroom)
    db.session.commit()

    flash(f'تم حذف الفصل "{classroom_name}" بنجاح', 'success')
    return redirect(url_for('admin.classrooms'))

@admin_bp.route('/classroom/<int:classroom_id>/view')
@login_required
@admin_required
def view_classroom(classroom_id):
    classroom = Classroom.query.get_or_404(classroom_id)
    
    # Get enrollments with students
    enrollments = classroom.enrollments
    
    # Get assignments with submission counts and additional info
    assignments = []
    for assignment in classroom.assignments:
        submissions = assignment.submissions if hasattr(assignment, 'submissions') else []
        assignment_dict = {
            'title': assignment.title,
            'due_date': assignment.due_date,
            'created_at': assignment.created_at,
            'submission_count': len(submissions),
            'total_marks': assignment.total_marks if hasattr(assignment, 'total_marks') else 0,
            'is_active': assignment.is_active if hasattr(assignment, 'is_active') else True
        }
        assignments.append(assignment_dict)

    # Get quizzes with additional info
    quiz_list = []
    if hasattr(classroom, 'quizzes'):
        for quiz in classroom.quizzes:
            quiz_dict = {
                'title': quiz.title,
                'end_time': quiz.end_time,
                'created_at': quiz.created_at,
                'is_active': quiz.is_active if hasattr(quiz, 'is_active') else True,
                'total_marks': quiz.total_marks if hasattr(quiz, 'total_marks') else 0
            }
            quiz_list.append(quiz_dict)

    # Get number of lessons
    lesson_count = len(classroom.lessons) if hasattr(classroom, 'lessons') else 0
    
    # Get teacher information
    teacher = classroom.teacher
    
    template = 'admin/admin-mobile/view_classroom.html' if is_mobile() else 'admin/view_classroom.html'

    # الحصول على قيم الألوان من إعدادات النظام
    primary_color = SystemSettings.get_setting('primary_color', '#3498db')  # اللون الافتراضي
    secondary_color = SystemSettings.get_setting('secondary_color', '#2ecc71')  # اللون الافتراضي

    
    return render_template(template, 
                           classroom=classroom,
                           primary_color=primary_color,
                           secondary_color=secondary_color,
                           enrollments=enrollments,
                           assignments=assignments,
                           quizzes=quiz_list,
                           teacher=teacher,
                           lesson_count=lesson_count,
                           students=enrollments,
                           now=datetime.utcnow())

@admin_bp.route('/classroom/<int:classroom_id>/toggle_status', methods=['POST'])
@login_required
@admin_required
def toggle_classroom_status(classroom_id):
    classroom = Classroom.query.get_or_404(classroom_id)

    classroom.is_active = not classroom.is_active
    status = "تفعيل" if classroom.is_active else "تعطيل"
    
    try:
        db.session.commit()
        flash(f'تم {status} الفصل "{classroom.name}" بنجاح', 'success')
    except Exception as e:
        db.session.rollback()
        flash('حدث خطأ أثناء تحديث حالة الفصل', 'error')
    
    return redirect(url_for('admin.classrooms'))

@admin_bp.route('/settings', methods=['GET', 'POST'])
@login_required
@admin_required
def settings():
    if request.method == 'POST':
        # General settings
        settings_map = {
            'system_name': ('general', request.form.get('system_name')),
            'system_description': ('general', request.form.get('system_description')),
            'default_language': ('general', request.form.get('default_language')),
            'default_currency': ('general', request.form.get('default_currency')),
            'trial_days': ('general', request.form.get('trial_days')),
            'max_file_size': ('general', request.form.get('max_file_size')),

            # Appearance settings
            'primary_color': ('appearance', request.form.get('primary_color')),
            'secondary_color': ('appearance', request.form.get('secondary_color')),
            'logo_url': ('appearance', request.form.get('logo_url')),

            # Contact settings
            'contact_email': ('contact', request.form.get('contact_email')),
            'contact_phone': ('contact', request.form.get('contact_phone'))
        }

        # Save all settings
        success = True
        for key, (group, value) in settings_map.items():
            if not SystemSettings.set_setting(key, value, group):
                success = False

        if success:
            flash('تم حفظ الإعدادات بنجاح', 'success')
        else:
            flash('حدث خطأ أثناء حفظ الإعدادات', 'danger')

        return redirect(url_for('admin.settings'))

    # Load current settings
    settings = {
        'system_name': SystemSettings.get_setting('system_name', 'الحصة'),
        'system_description': SystemSettings.get_setting('system_description', 'منصة تعليمية متكاملة لإدارة العملية التعليمية بكفاءة'),
        'default_language': SystemSettings.get_setting('default_language', 'ar'),
        'default_currency': SystemSettings.get_setting('default_currency', 'SAR'),
        'trial_days': SystemSettings.get_setting('trial_days', '14'),
        'max_file_size': SystemSettings.get_setting('max_file_size', '16'),
        'primary_color': SystemSettings.get_setting('primary_color', '#4e73df'),
        'secondary_color': SystemSettings.get_setting('secondary_color', '#1cc88a'),
        'logo_url': SystemSettings.get_setting('logo_url', '/static/img/logo.png'),
        'contact_email': SystemSettings.get_setting('contact_email', 'support@alhesa.com'),
        'contact_phone': SystemSettings.get_setting('contact_phone', '+966 55 555 5555')
    }

    now = datetime.utcnow()

    template = 'admin/admin-mobile/settings.html' if is_mobile() else 'admin/settings.html'

    # الحصول على قيم الألوان من إعدادات النظام
    primary_color = SystemSettings.get_setting('primary_color', '#3498db')  # اللون الافتراضي
    secondary_color = SystemSettings.get_setting('secondary_color', '#2ecc71')  # اللون الافتراضي


    return render_template(template, 
                           settings=settings, 
                           now=now,
                           primary_color=primary_color,
                           secondary_color=secondary_color,)

@admin_bp.route('/payment/<int:payment_id>/approve', methods=['POST'])
@login_required
@admin_required
def approve_payment(payment_id):
    """
    قبول دفعة الطالب في الفصل (للإدارة)
    """
    payment = Payment.query.get_or_404(payment_id)
    
    if payment.status != 'pending':
        flash('لا يمكن قبول هذا الطلب لأنه تم معالجته مسبقاً', 'warning')
        return redirect(url_for('admin.subscriptions'))
    
    # تحديث حالة الدفع
    payment.status = 'success'
    
    # تفعيل اشتراك الطالب في الفصل
    enrollment = ClassroomEnrollment.query.filter_by(
        user_id=payment.user_id,
        classroom_id=payment.classroom_id
    ).first()
    
    if enrollment:
        enrollment.payment_status = 'paid'
        enrollment.is_active = True
        enrollment.payment_date = datetime.utcnow()
    
    db.session.commit()
    
    flash('تم تفعيل اشتراك الطالب بنجاح', 'success')
    return redirect(url_for('admin.subscriptions'))

@admin_bp.route('/payment/<int:payment_id>/reject', methods=['POST'])
@login_required
@admin_required
def reject_payment(payment_id):
    """
    رفض دفعة الطالب في الفصل (للإدارة)
    """
    payment = Payment.query.get_or_404(payment_id)
    
    if payment.status != 'pending':
        flash('لا يمكن رفض هذا الطلب لأنه تم معالجته مسبقاً', 'warning')
        return redirect(url_for('admin.subscriptions'))
    
    # تحديث حالة الدفع
    payment.status = 'failed'
    
    # تعطيل اشتراك الطالب في الفصل
    enrollment = ClassroomEnrollment.query.filter_by(
        user_id=payment.user_id,
        classroom_id=payment.classroom_id
    ).first()
    
    if enrollment:
        enrollment.payment_status = 'failed'
        enrollment.is_active = False
    
    db.session.commit()
    
    flash('تم رفض عملية الدفع', 'danger')
    return redirect(url_for('admin.subscriptions'))

@admin_bp.route('/users/export')
@login_required
@admin_required
def export_users():
    """تصدير بيانات المستخدمين إلى ملف Excel"""
    try:
        # Get filter parameters
        role_filter = request.args.get('role', '')
        status_filter = request.args.get('status', '')
        search = request.args.get('search', '')

        # Base query
        query = User.query

        # Apply filters
        if role_filter and role_filter in [Role.STUDENT, Role.TEACHER, Role.ASSISTANT, Role.ADMIN]:
            query = query.filter_by(role=role_filter)

        if status_filter:
            if status_filter == 'active':
                query = query.filter_by(is_active=True)
            elif status_filter == 'inactive':
                query = query.filter_by(is_active=False)

        if search:
            query = query.filter(
                (User.name.ilike(f'%{search}%')) | 
                (User.phone.ilike(f'%{search}%')) |
                (User.email.ilike(f'%{search}%'))
            )

        # Get all users
        users = query.order_by(User.created_at.desc()).all()

        # Prepare data for export
        users_data = []
        for user in users:
            role_name = {
                'student': 'طالب',
                'teacher': 'معلم', 
                'assistant': 'مساعد',
                'admin': 'مسؤول'
            }.get(user.role, user.role)
            
            status_name = 'نشط' if user.is_active else 'معطل'
            
            user_data = {
                'name': user.name,
                'phone': user.phone,
                'email': user.email or '',
                'role': role_name,
                'created_at': user.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                'status': status_name,
                'updated_at': user.updated_at.strftime('%Y-%m-%d %H:%M:%S') if user.updated_at else '',
                'address': user.address or ''
            }
            users_data.append(user_data)

        return jsonify({
            'success': True,
            'data': users_data,
            'total': len(users_data)
        })

    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@admin_bp.route('/users/export/excel')
@login_required
@admin_required  
def export_users_excel():
    """تصدير بيانات المستخدمين إلى ملف Excel"""
    try:
        # استيراد مكتبة openpyxl
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            excel_available = True
        except ImportError:
            return jsonify({
                'success': False,
                'error': 'مكتبة openpyxl غير مثبتة'
            }), 500

        # الحصول على جميع المستخدمين
        users = User.query.all()
        
        # إحصائيات
        total_users = len(users)
        active_users = len([u for u in users if u.is_active])
        inactive_users = total_users - active_users
        
        # إحصائيات الأدوار
        students_count = len([u for u in users if u.role == 'student'])
        teachers_count = len([u for u in users if u.role == 'teacher'])
        assistants_count = len([u for u in users if u.role == 'assistant'])
        admins_count = len([u for u in users if u.role == 'admin'])

        # إنشاء كتاب Excel جديد
        wb = Workbook()
        ws = wb.active
        ws.title = "بيانات المستخدمين"
        
        # تعيين خصائص الكتاب
        wb.properties.creator = "AL-7esa Platform"
        wb.properties.title = "تقرير المستخدمين"
        
        # تعريف الأنماط البسيطة
        header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        normal_font = Font(name='Arial', size=11)
        primary_fill = PatternFill(start_color='2E4057', end_color='2E4057', fill_type='solid')
        success_fill = PatternFill(start_color='D4F6D4', end_color='D4F6D4', fill_type='solid')
        warning_fill = PatternFill(start_color='FFF3CD', end_color='FFF3CD', fill_type='solid')
        
        # المحاذاة
        center_alignment = Alignment(horizontal='center', vertical='center')
        right_alignment = Alignment(horizontal='right', vertical='center')
        
        # الحدود
        thin_border = Side(border_style="thin", color="B0B0B0")
        border = Border(left=thin_border, right=thin_border, top=thin_border, bottom=thin_border)

        # العنوان الرئيسي
        ws.merge_cells('A1:I2')
        title_cell = ws['A1']
        title_cell.value = 'تقرير المستخدمين - منصة الحصة التعليمية'
        title_cell.font = Font(name='Arial', size=16, bold=True, color='FFFFFF')
        title_cell.alignment = center_alignment
        title_cell.fill = primary_fill
        
        # معلومات التقرير
        from datetime import datetime
        current_time = datetime.now()
        
        info_row = 3
        ws[f'A{info_row}'] = f'تاريخ التقرير: {current_time.strftime("%Y-%m-%d %H:%M")}'
        ws[f'A{info_row}'].font = normal_font
        
        info_row += 1
        ws[f'A{info_row}'] = f'إجمالي المستخدمين: {total_users} | النشطين: {active_users} | المعطلين: {inactive_users}'
        ws[f'A{info_row}'].font = normal_font
        
        info_row += 1
        ws[f'A{info_row}'] = f'الطلاب: {students_count} | المعلمين: {teachers_count} | المساعدين: {assistants_count} | المديرين: {admins_count}'
        ws[f'A{info_row}'].font = normal_font
        
        # عناوين الأعمدة
        headers = [
            'الرقم',
            'الاسم الكامل',
            'رقم الهاتف', 
            'البريد الإلكتروني',
            'الدور',
            'الحالة',
            'تاريخ التسجيل',
            'آخر تحديث',
            'العنوان'
        ]

        header_row = 7
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col, value=header)
            cell.font = header_font
            cell.fill = primary_fill
            cell.alignment = center_alignment
            cell.border = border

        # إضافة بيانات المستخدمين
        for row, user in enumerate(users, header_row + 1):
            role_name = {
                'student': 'طالب',
                'teacher': 'معلم', 
                'assistant': 'مساعد',
                'admin': 'مدير'
            }.get(user.role, user.role)
            
            status_name = 'نشط' if user.is_active else 'معطل'
            
            user_data = [
                row - header_row,  # الرقم
                user.name,
                user.phone,
                user.email or '',
                role_name,
                status_name,
                user.created_at.strftime('%Y-%m-%d'),
                user.updated_at.strftime('%Y-%m-%d') if user.updated_at else '',
                user.address or ''
            ]
            
            for col, value in enumerate(user_data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.font = normal_font
                cell.alignment = center_alignment if col == 1 else right_alignment
                cell.border = border
                
                # تلوين الصفوف بالتناوب
                if row % 2 == 0:
                    cell.fill = PatternFill(start_color='F8F9FA', end_color='F8F9FA', fill_type='solid')
                
                # تلوين خاص للحالة
                if col == 6:  # عمود الحالة
                    if user.is_active:
                        cell.fill = success_fill
                        cell.font = Font(name='Arial', size=11, bold=True, color='155724')
                    else:
                        cell.fill = warning_fill
                        cell.font = Font(name='Arial', size=11, bold=True, color='856404')

        # تعديل عرض الأعمدة
        column_widths = [8, 25, 15, 30, 12, 10, 15, 15, 35]
        for col, width in enumerate(column_widths, 1):
            column_letter = get_column_letter(col)
            ws.column_dimensions[column_letter].width = width

        # حفظ الملف
        from io import BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        from flask import send_file
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'تقرير_المستخدمين_{timestamp}.xlsx'

        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@admin_bp.route('/classrooms/export/excel')
@login_required
@admin_required
def export_classrooms_excel():
    """تصدير بيانات الفصول الدراسية إلى ملف Excel مع تصميم مميز"""
    try:
        # استيراد مكتبة openpyxl
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            from openpyxl.drawing.image import Image
            excel_available = True
        except ImportError:
            excel_available = False

        if not excel_available:
            return jsonify({
                'success': False,
                'error': 'مكتبة Excel غير متاحة. يرجى تثبيت openpyxl'
            }), 400

        # Get filter parameters
        is_free_filter = request.args.get('is_free', '')
        subject_filter = request.args.get('subject', '')
        grade_filter = request.args.get('grade', '')
        sort_filter = request.args.get('sort', '')

        # Base query with joins
        query = Classroom.query.options(
            db.joinedload(Classroom.teacher),
            db.joinedload(Classroom.assistant),
            db.joinedload(Classroom.enrollments)
        )

        # Apply filters
        if is_free_filter == 'true':
            query = query.filter_by(is_free=True)
        elif is_free_filter == 'false':
            query = query.filter_by(is_free=False)

        if subject_filter:
            query = query.filter(Classroom.subject.ilike(f'%{subject_filter}%'))

        if grade_filter:
            query = query.filter(Classroom.grade.ilike(f'%{grade_filter}%'))

        # Apply sorting
        if sort_filter == 'newest':
            query = query.order_by(Classroom.created_at.desc())
        elif sort_filter == 'oldest':
            query = query.order_by(Classroom.created_at.asc())
        else:
            query = query.order_by(Classroom.created_at.desc())

        # Get all classrooms
        classrooms = query.all()

        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "تقرير الفصول الدراسية"

        # Define colors and styles
        header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")  # Dark Blue
        subheader_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")  # Blue
        free_fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")  # Green
        paid_fill = PatternFill(start_color="F59E0B", end_color="F59E0B", fill_type="solid")  # Orange
        
        header_font = Font(bold=True, color="FFFFFF", size=14)
        subheader_font = Font(bold=True, color="FFFFFF", size=12)
        title_font = Font(bold=True, color="1E3A8A", size=18)
        normal_font = Font(color="374151", size=11)
        
        center_alignment = Alignment(horizontal="center", vertical="center")
        left_alignment = Alignment(horizontal="left", vertical="center")
        
        thin_border = Border(
            left=Side(style='thin', color='E5E7EB'),
            right=Side(style='thin', color='E5E7EB'),
            top=Side(style='thin', color='E5E7EB'),
            bottom=Side(style='thin', color='E5E7EB')
        )

        # Add title and summary section
        ws.merge_cells('A1:I3')
        title_cell = ws['A1']
        title_cell.value = "تقرير الفصول الدراسية - منصة الحصة التعليمية"
        title_cell.font = title_font
        title_cell.alignment = center_alignment
        title_cell.fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")

        # Add date and statistics
        from datetime import datetime
        current_date = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        ws.merge_cells('A4:I4')
        date_cell = ws['A4']
        date_cell.value = f"تاريخ التقرير: {current_date}"
        date_cell.font = Font(color="6B7280", size=12, italic=True)
        date_cell.alignment = center_alignment

        # Statistics row
        total_classrooms = len(classrooms)
        free_classrooms = len([c for c in classrooms if c.is_free])
        paid_classrooms = total_classrooms - free_classrooms
        total_students = sum(len(c.enrollments) for c in classrooms)

        ws.merge_cells('A6:B6')
        ws['A6'] = "إجمالي الفصول:"
        ws['A6'].font = Font(bold=True, color="1E3A8A")
        ws['C6'] = total_classrooms
        ws['C6'].font = Font(bold=True, color="059669")

        ws.merge_cells('D6:E6')
        ws['D6'] = "الفصول المجانية:"
        ws['D6'].font = Font(bold=True, color="1E3A8A")
        ws['F6'] = free_classrooms
        ws['F6'].font = Font(bold=True, color="059669")

        ws.merge_cells('G6:H6')
        ws['G6'] = "الفصول المدفوعة:"
        ws['G6'].font = Font(bold=True, color="1E3A8A")
        ws['I6'] = paid_classrooms
        ws['I6'].font = Font(bold=True, color="D97706")

        ws.merge_cells('A7:B7')
        ws['A7'] = "إجمالي الطلاب:"
        ws['A7'].font = Font(bold=True, color="1E3A8A")
        ws['C7'] = total_students
        ws['C7'].font = Font(bold=True, color="7C2D12")

        # Headers for data table
        headers = [
            'كود الفصل',
            'اسم الفصل', 
            'المادة',
            'المستوى/الصف',
            'المعلم',
            'المساعد',
            'عدد الطلاب',
            'النوع',
            'السعر (جنية)',
            'تاريخ الإنشاء'
        ]

        # Add headers starting from row 9
        start_row = 9
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=start_row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = thin_border

        # Add classroom data
        for row_idx, classroom in enumerate(classrooms, start_row + 1):
            # Basic information
            ws.cell(row=row_idx, column=1, value=classroom.code)
            ws.cell(row=row_idx, column=2, value=classroom.name)
            ws.cell(row=row_idx, column=3, value=classroom.subject)
            ws.cell(row=row_idx, column=4, value=classroom.grade)
            ws.cell(row=row_idx, column=5, value=classroom.teacher.name if classroom.teacher else 'غير محدد')
            ws.cell(row=row_idx, column=6, value=classroom.assistant.name if classroom.assistant else '-')
            ws.cell(row=row_idx, column=7, value=len(classroom.enrollments))
            
            # Type and price with color coding
            type_cell = ws.cell(row=row_idx, column=8)
            price_cell = ws.cell(row=row_idx, column=9)
            
            if classroom.is_free:
                type_cell.value = "مجاني"
                type_cell.fill = free_fill
                type_cell.font = Font(bold=True, color="FFFFFF")
                price_cell.value = "مجاني"
                price_cell.fill = free_fill
                price_cell.font = Font(bold=True, color="FFFFFF")
            else:
                type_cell.value = "مدفوع"
                type_cell.fill = paid_fill
                type_cell.font = Font(bold=True, color="FFFFFF")
                price_cell.value = f"{classroom.price or 0} جنية"
                price_cell.fill = paid_fill
                price_cell.font = Font(bold=True, color="FFFFFF")
            
            # Creation date
            date_cell = ws.cell(row=row_idx, column=10, value=classroom.created_at.strftime('%Y-%m-%d'))
            
            # Apply styling to all cells in the row
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=row_idx, column=col)
                cell.font = normal_font
                cell.alignment = center_alignment
                cell.border = thin_border
                
                # Alternate row colors
                if row_idx % 2 == 0:
                    if col not in [8, 9]:  # Skip type and price columns
                        cell.fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")

        # Auto-adjust column widths
        for col in range(1, len(headers) + 1):
            column_letter = get_column_letter(col)
            max_length = 0
            
            # Check header length
            max_length = max(max_length, len(str(headers[col-1])))
            
            # Check data length
            for row in range(start_row + 1, ws.max_row + 1):
                cell_value = str(ws[f"{column_letter}{row}"].value or "")
                max_length = max(max_length, len(cell_value))
            
            # Set column width with some padding
            adjusted_width = min(max_length + 3, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Add a summary section at the bottom
        summary_row = ws.max_row + 3
        
        ws.merge_cells(f'A{summary_row}:J{summary_row}')
        summary_cell = ws[f'A{summary_row}']
        summary_cell.value = "ملخص التقرير"
        summary_cell.font = Font(bold=True, color="FFFFFF", size=14)
        summary_cell.fill = header_fill
        summary_cell.alignment = center_alignment

        # Summary statistics
        summary_data = [
            ("إجمالي الفصول الدراسية:", total_classrooms),
            ("الفصول المجانية:", free_classrooms),
            ("الفصول المدفوعة:", paid_classrooms),
            ("إجمالي الطلاب المسجلين:", total_students),
            ("متوسط الطلاب لكل فصل:", round(total_students / total_classrooms, 2) if total_classrooms > 0 else 0)
        ]

        for i, (label, value) in enumerate(summary_data):
            row = summary_row + 1 + i
            ws.cell(row=row, column=1, value=label).font = Font(bold=True, color="1E3A8A")
            ws.cell(row=row, column=2, value=value).font = Font(bold=True, color="059669")

        # Create response
        from io import BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        from flask import send_file
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'تقرير_الفصول_الدراسية_{timestamp}.xlsx'

        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        current_app.logger.error(f"خطأ في تصدير الفصول: {str(e)}")
        return jsonify({
            'success': False,
            'error': f'حدث خطأ أثناء تصدير البيانات: {str(e)}'
        }), 500

@admin_bp.route('/backup', methods=['POST'])
@login_required
@admin_required
def backup_database():
    """
    إنشاء نسخة احتياطية من قاعدة البيانات وتنزيلها
    """
    return _create_database_backup()

@admin_bp.route('/backup/download', methods=['GET'])
@login_required
@admin_required
def backup_database_get():
    """
    إنشاء نسخة احتياطية من قاعدة البيانات وتنزيلها (GET request)
    """
    return _create_database_backup()

def _create_database_backup():
    """
    دالة مساعدة لإنشاء النسخة الاحتياطية - تنزيل ملف قاعدة البيانات الأصلي
    """
    try:
        import os
        import shutil
        import tempfile
        from datetime import datetime
        from flask import send_file, current_app, flash, redirect, url_for
        
        # المسار المحدد الذي ذكره المستخدم
        specific_db_path = r'C:\Users\Hamada Salim G Trd\Desktop\AL-7esa\instance\al-7esa.db'
        
        # البحث عن ملف قاعدة البيانات في عدة مواقع محتملة
        possible_db_paths = [
            # المسار المحدد من المستخدم أولاً
            specific_db_path,
            # في مجلد instance
            os.path.join(current_app.instance_path, 'al-7esa.db'),
            # في المجلد الجذر للمشروع
            os.path.join(os.path.dirname(current_app.root_path), 'al-7esa.db'),
            os.path.join(current_app.root_path, 'al-7esa.db'),
            # في المجلد الحالي
            os.path.join(os.getcwd(), 'al-7esa.db'),
            os.path.join(os.getcwd(), 'instance', 'al-7esa.db'),
        ]
        
        # إضافة مسار من التكوين إذا كان موجوداً
        db_uri = current_app.config.get('SQLALCHEMY_DATABASE_URI', '')
        if db_uri.startswith('sqlite:///'):
            config_db_path = db_uri.replace('sqlite:///', '')
            if not os.path.isabs(config_db_path):
                config_db_path = os.path.join(current_app.root_path, config_db_path)
            possible_db_paths.insert(1, config_db_path)  # إدراجه بعد المسار المحدد
        
        db_path = None
        for path in possible_db_paths:
            if path and os.path.exists(path) and os.path.getsize(path) > 0:
                db_path = path
                current_app.logger.info(f"تم العثور على قاعدة البيانات في: {db_path}")
                break
        
        if not db_path:
            # محاولة البحث عن أي ملف .db في المشروع
            search_paths = [current_app.root_path, os.getcwd()]
            for search_root in search_paths:
                for root, dirs, files in os.walk(search_root):
                    for file in files:
                        if file.endswith('.db') and os.path.getsize(os.path.join(root, file)) > 0:
                            db_path = os.path.join(root, file)
                            current_app.logger.info(f"تم العثور على قاعدة بيانات في: {db_path}")
                            break
                    if db_path:
                        break
                if db_path:
                    break
        
        if not db_path:
            current_app.logger.error("لم يتم العثور على ملف قاعدة البيانات")
            flash('ملف قاعدة البيانات غير موجود أو فارغ', 'error')
            return redirect(url_for('admin.dashboard'))
        
        # التحقق من حجم الملف
        db_size = os.path.getsize(db_path)
        if db_size == 0:
            flash('ملف قاعدة البيانات فارغ', 'error')
            return redirect(url_for('admin.dashboard'))
        
        current_app.logger.info(f"حجم قاعدة البيانات: {db_size} بايت - المسار: {db_path}")
        
        # إنشاء اسم الملف مع التاريخ والوقت
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        backup_filename = f'al-7esa_backup_{timestamp}.db'
        
        # إنشاء نسخة من ملف قاعدة البيانات الأصلي
        try:
            # إنشاء ملف مؤقت
            temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.db')
            temp_file.close()
            
            # نسخ ملف قاعدة البيانات إلى الملف المؤقت
            shutil.copy2(db_path, temp_file.name)
            
            # تسجيل عملية النسخ الاحتياطي
            current_app.logger.info(f"تم إنشاء نسخة احتياطية من ملف .db بواسطة المستخدم: {current_user.name} - حجم الملف: {db_size} بايت - المسار المصدر: {db_path}")
            
            # إرسال ملف قاعدة البيانات الأصلي للتنزيل
            def remove_temp_file(response):
                try:
                    os.unlink(temp_file.name)
                except Exception:
                    pass
                return response
            
            response = send_file(
                temp_file.name,
                as_attachment=True,
                download_name=backup_filename,
                mimetype='application/octet-stream'
            )
            
            # إضافة callback لحذف الملف المؤقت بعد الإرسال
            response.call_on_close(lambda: os.unlink(temp_file.name) if os.path.exists(temp_file.name) else None)
            
            return response
            
        except Exception as e:
            current_app.logger.error(f"خطأ في إنشاء النسخة الاحتياطية: {str(e)}")
            flash(f'خطأ في إنشاء ملف النسخة الاحتياطية: {str(e)}', 'error')
            return redirect(url_for('admin.dashboard'))
        
    except Exception as e:
        current_app.logger.error(f"خطأ في إنشاء النسخة الاحتياطية: {str(e)}")
        flash(f'حدث خطأ أثناء إنشاء النسخة الاحتياطية: {str(e)}', 'error')
        return redirect(url_for('admin.dashboard'))

@admin_bp.route('/reset_password/<int:user_id>', methods=['POST'])
@login_required
@admin_required
def reset_password(user_id):
    """
    إعادة تعيين كلمة مرور المستخدم إلى كلمة المرور الافتراضية
    """
    try:
        # العثور على المستخدم
        user = User.query.get_or_404(user_id)
        
        # تعيين كلمة المرور الافتراضية
        default_password = "12345678"
        user.set_password(default_password)
        
        # حفظ التغييرات
        db.session.commit()
        
        flash(f'تم إعادة تعيين كلمة مرور المستخدم {user.username} بنجاح إلى كلمة المرور الافتراضية (12345678)', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'خطأ في إعادة تعيين كلمة المرور: {str(e)}', 'error')
    
    return redirect(url_for('admin.users'))


# =================== مسارات إدارة البانرات ===================

@admin_bp.route('/banners')
@login_required
@admin_required
def banners():
    """صفحة إدارة البانرات"""
    page = request.args.get('page', 1, type=int)
    banners = Banner.query.order_by(Banner.priority.desc(), Banner.created_at.desc()).paginate(
        page=page, per_page=20
    )
    
    template = 'admin/admin-mobile/banners.html' if is_mobile() else 'admin/banners.html'
    
    # الحصول على قيم الألوان من إعدادات النظام
    primary_color = SystemSettings.get_setting('primary_color', '#3498db')
    secondary_color = SystemSettings.get_setting('secondary_color', '#2ecc71')
    
    return render_template(template, 
                         banners=banners,
                         primary_color=primary_color,
                         secondary_color=secondary_color)

@admin_bp.route('/banner/new', methods=['GET', 'POST'])
@login_required
@admin_required
def new_banner():
    """إضافة بانر جديد"""
    if request.method == 'POST':
        try:
            title = request.form.get('title')
            description = request.form.get('description')
            link_url = request.form.get('link_url')
            target_roles = request.form.get('target_roles', 'all')
            priority = int(request.form.get('priority', 0))
            start_date = request.form.get('start_date')
            end_date = request.form.get('end_date')
            
            # معالجة رفع الصورة
            image_url = None
            if 'banner_image' in request.files:
                file = request.files['banner_image']
                if file and file.filename:
                    from werkzeug.utils import secure_filename
                    import uuid
                    
                    # التحقق من نوع الملف
                    allowed_extensions = {'png', 'jpg', 'jpeg', 'gif', 'webp'}
                    if '.' in file.filename and file.filename.rsplit('.', 1)[1].lower() in allowed_extensions:
                        # إنشاء اسم ملف فريد
                        unique_filename = f"{uuid.uuid4().hex}_{secure_filename(file.filename)}"
                        
                        # إنشاء مجلد البنرات
                        upload_folder = os.path.join(current_app.root_path, 'static', 'uploads', 'banners')
                        os.makedirs(upload_folder, exist_ok=True)
                        
                        # حفظ الملف
                        filepath = os.path.join(upload_folder, unique_filename)
                        file.save(filepath)
                        
                        # تعيين رابط الصورة
                        image_url = f"/static/uploads/banners/{unique_filename}"
                    else:
                        flash('نوع الملف غير مدعوم. يرجى اختيار صورة بصيغة PNG, JPG, JPEG, GIF, أو WEBP', 'error')
                        return redirect(request.url)
            
            if not image_url:
                flash('يرجى اختيار صورة للبنر', 'error')
                return redirect(request.url)
            
            # تحويل التواريخ
            start_date_obj = datetime.strptime(start_date, '%Y-%m-%d') if start_date else None
            end_date_obj = datetime.strptime(end_date, '%Y-%m-%d') if end_date else None
            
            new_banner = Banner(
                title=title,
                description=description,
                image_url=image_url,
                link_url=link_url,
                target_roles=target_roles,
                priority=priority,
                start_date=start_date_obj,
                end_date=end_date_obj,
                created_by=current_user.id
            )
            
            db.session.add(new_banner)
            db.session.commit()
            
            flash('تم إضافة البانر بنجاح', 'success')
            return redirect(url_for('admin.banners'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'خطأ في إضافة البانر: {str(e)}', 'error')
    
    template = 'admin/admin-mobile/edit_banner.html' if is_mobile() else 'admin/edit_banner.html'
    
    # الحصول على قيم الألوان من إعدادات النظام
    primary_color = SystemSettings.get_setting('primary_color', '#3498db')
    secondary_color = SystemSettings.get_setting('secondary_color', '#2ecc71')
    
    return render_template(template,
                         banner=None,
                         primary_color=primary_color,
                         secondary_color=secondary_color)

@admin_bp.route('/banner/<int:banner_id>/edit', methods=['GET', 'POST'])
@login_required
@admin_required
def edit_banner(banner_id):
    """تعديل بانر"""
    banner = Banner.query.get_or_404(banner_id)
    
    if request.method == 'POST':
        try:
            banner.title = request.form.get('title')
            banner.description = request.form.get('description')
            banner.link_url = request.form.get('link_url')
            banner.target_roles = request.form.get('target_roles', 'all')
            banner.priority = int(request.form.get('priority', 0))
            
            # معالجة رفع صورة جديدة
            if 'banner_image' in request.files:
                file = request.files['banner_image']
                if file and file.filename:
                    from werkzeug.utils import secure_filename
                    import uuid
                    
                    # التحقق من نوع الملف
                    allowed_extensions = {'png', 'jpg', 'jpeg', 'gif', 'webp'}
                    if '.' in file.filename and file.filename.rsplit('.', 1)[1].lower() in allowed_extensions:
                        # حذف الصورة القديمة إذا كانت موجودة
                        if banner.image_url and banner.image_url.startswith('/static/uploads/banners/'):
                            old_filepath = os.path.join(current_app.root_path, banner.image_url.lstrip('/'))
                            if os.path.exists(old_filepath):
                                try:
                                    os.remove(old_filepath)
                                except:
                                    pass  # لا بأس إذا لم نتمكن من حذف الملف القديم
                        
                        # إنشاء اسم ملف فريد
                        unique_filename = f"{uuid.uuid4().hex}_{secure_filename(file.filename)}"
                        
                        # إنشاء مجلد البنرات
                        upload_folder = os.path.join(current_app.root_path, 'static', 'uploads', 'banners')
                        os.makedirs(upload_folder, exist_ok=True)
                        
                        # حفظ الملف
                        filepath = os.path.join(upload_folder, unique_filename)
                        file.save(filepath)
                        
                        # تحديث رابط الصورة
                        banner.image_url = f"/static/uploads/banners/{unique_filename}"
                    else:
                        flash('نوع الملف غير مدعوم. يرجى اختيار صورة بصيغة PNG, JPG, JPEG, GIF, أو WEBP', 'error')
                        return redirect(request.url)
            
            start_date = request.form.get('start_date')
            end_date = request.form.get('end_date')
            
            banner.start_date = datetime.strptime(start_date, '%Y-%m-%d') if start_date else None
            banner.end_date = datetime.strptime(end_date, '%Y-%m-%d') if end_date else None
            
            # تحديث حالة التفعيل
            banner.is_active = 'is_active' in request.form
            
            db.session.commit()
            
            flash('تم تحديث البانر بنجاح', 'success')
            return redirect(url_for('admin.banners'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'خطأ في تحديث البانر: {str(e)}', 'error')
    
    template = 'admin/admin-mobile/edit_banner.html' if is_mobile() else 'admin/edit_banner.html'
    
    # الحصول على قيم الألوان من إعدادات النظام
    primary_color = SystemSettings.get_setting('primary_color', '#3498db')
    secondary_color = SystemSettings.get_setting('secondary_color', '#2ecc71')
    
    return render_template(template,
                         banner=banner,
                         primary_color=primary_color,
                         secondary_color=secondary_color)

@admin_bp.route('/banner/<int:banner_id>/toggle', methods=['POST'])
@login_required
@admin_required
def toggle_banner(banner_id):
    """تفعيل/إلغاء تفعيل بانر"""
    banner = Banner.query.get_or_404(banner_id)
    banner.is_active = not banner.is_active
    db.session.commit()
    
    status = 'تم تفعيل' if banner.is_active else 'تم إلغاء تفعيل'
    flash(f'{status} البانر بنجاح', 'success')
    return redirect(url_for('admin.banners'))

@admin_bp.route('/banner/<int:banner_id>/delete', methods=['POST'])
@login_required
@admin_required
def delete_banner(banner_id):
    """حذف بانر"""
    banner = Banner.query.get_or_404(banner_id)
    
    try:
        db.session.delete(banner)
        db.session.commit()
        flash('تم حذف البانر بنجاح', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'خطأ في حذف البانر: {str(e)}', 'error')
    
    return redirect(url_for('admin.banners'))
